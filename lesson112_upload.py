# type: ignore

import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions

file_path = "C:\\Users\\vinic\\Downloads\\download.xlsx"
fruit = 'Apple'
new_value = '1000'

driver = webdriver.Chrome()
driver.get('https://rahulshettyacademy.com/upload-download-test/')
driver.implicitly_wait(5)


driver.find_element(By.CSS_SELECTOR, '#downloadButton').click()
time.sleep(1)
######
book = openpyxl.load_workbook(file_path)
sheet = book.active

for column in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=column).value == 'fruit_name':
        fruit_column = column
        for row in range(1, sheet.max_column+1):
            if sheet.cell(row=row, column=column).value == fruit:
                fruit_row = row

for column in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=column).value == 'price':
        price_column = column

price_cell = sheet.cell(row=fruit_row, column=price_column)
price_cell.value = new_value
book.save(file_path)
######

file_input = driver.find_element(By.CSS_SELECTOR, 'input[type="file"]')
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(
    By.XPATH, '//div[text()="Price"]').get_attribute('data-column-id')
price_locator = f'//div[text()="{fruit}"]/parent::div/parent::div/'\
    f'div[@id="cell-{price_column}-undefined"]'
actual_price = driver.find_element(By.XPATH, price_locator).text
print(actual_price)
assert actual_price == new_value
