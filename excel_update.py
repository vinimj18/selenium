import openpyxl

file_path = "C:\\Users\\vinic\\Downloads\\download.xlsx"

book = openpyxl.load_workbook(file_path)
sheet = book.active
fruit = "Apple"
book = openpyxl.load_workbook(file_path)
sheet = book.active

for column in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=column).value == 'fruit_name':
        fruit_column = column
        for row in range(1, sheet.max_column+1):
            if sheet.cell(row=row, column=column).value == fruit:
                fruit_row = row

for column in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=column).value == 'price':
        price_column = column

price_cell = sheet.cell(row=fruit_row, column=price_column)
price_cell.value = '500'
book.save(file_path)
